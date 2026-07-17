import streamlit as st
import pandas as pd
from datetime import datetime, time
import openpyxl
import os

st.set_page_config(page_title="Saisie Heures Supplémentaires", layout="wide")

st.title("🚓 Saisie des Heures Supplémentaires")
st.subheader("Interface mobile pour la Police Municipale")

# Vérification de la présence de la matrice
TEMPLATE_NAME = "BOTTE SUP.xlsx"
if not os.path.exists(TEMPLATE_NAME):
    st.error(f"⚠️ Le fichier gabarit '{TEMPLATE_NAME}' est introuvable dans le dossier GitHub. Veuillez l'ajouter.")

# --- EN-TÊTE DE LA DEMANDE ---
st.sidebar.header("👤 Agent & Période")
agent_nom = st.sidebar.text_input("Nom et Prénom de l'agent :", "BOTTE Philippe")
service_nom = st.sidebar.text_input("Service / Direction :", "Police Municipale")
mois_annee = st.sidebar.text_input("Mois / Année (ex: JUILLET 2026) :", "JUILLET 2026")

# --- MÉMOIRE DE LA SESSION ---
if 'registre_heures' not in st.session_state:
    st.session_state.registre_heures = []

# --- FORMULAIRE DE SAISIE ---
st.write("### 📅 Saisie d'une vacation")
col1, col2 = st.columns(2)

with col1:
    date_v = st.date_input("Date de l'effet :", datetime.now())
    motif_v = st.selectbox("Motif de la mission :", [
        "Main courantes et consignes",
        "Renfort",
        "Intervention extérieure",
        "Autre"
    ])
    if motif_v == "Autre":
        motif_v = st.text_input("Spécifiez le motif :", "")

with col2:
    h_deb = st.time_input("Heure de Début :", time(23, 0))
    h_fin = st.time_input("Heure de Fin :", time(3, 0))

if st.button("➕ Enregistrer cette vacation", use_container_width=True):
    # Calcul d'amplitude avec passage de minuit (ex: 23h à 3h du matin)
    dt_deb = datetime.combine(date_v, h_deb)
    dt_fin = datetime.combine(date_v, h_fin)
    if dt_fin <= dt_deb:
        heures_totales = ((dt_fin - dt_deb).total_seconds() / 3600) + 24
    else:
        heures_totales = (dt_fin - dt_deb).total_seconds() / 3600

    # Ventilation automatique
    h_sem, h_dim, h_nuit = 0.0, 0.0, 0.0
    est_dimanche = date_v.weekday() == 6

    if h_deb.hour >= 22 or h_deb.hour < 5:
        h_nuit = heures_totales
    elif est_dimanche:
        h_dim = heures_totales
    else:
        h_sem = heures_totales

    st.session_state.registre_heures.append({
        "Date": date_v.strftime("%Y-%m-%d"),
        "Horaire": f"{h_deb.strftime('%Hh%M')} à {h_fin.strftime('%Hh%M')}",
        "Semaine": h_sem if h_sem > 0 else None,
        "Dimanche": h_dim if h_dim > 0 else None,
        "Nuit": h_nuit if h_nuit > 0 else None,
        "Motif": motif_v
    })
    st.success("Vacation ajoutée au tableau mensuel.")

# --- AFFICHAGE DU TABLEAU ET EXPORT ---
if st.session_state.registre_heures:
    st.markdown("---")
    st.write("### 📋 Récapitulatif avant export")
    df_visu = pd.DataFrame(st.session_state.registre_heures).fillna("")
    st.dataframe(df_visu, use_container_width=True)

    if st.button("🗑️ Réinitialiser tout le tableau", type="secondary"):
        st.session_state.registre_heures = []
        st.rerun()

    st.markdown("---")
    if st.button("📥 Générer le fichier Excel officiel", type="primary", use_container_width=True):
        try:
            wb = openpyxl.load_workbook(TEMPLATE_NAME)
            ws = wb["AUTRE SERVICE"]

            # Remplissage de l'en-tête officiel
            ws["C5"] = agent_nom
            ws["F5"] = f"Mois :  {mois_annee}"
            ws["C7"] = service_nom

            # Nettoyage des anciennes lignes types
            for r in range(12, 26):
                for c in range(1, 7):
                    ws.cell(row=r, column=c, value=None)

            # Injection des nouvelles données
            start_row = 12
            for idx, data in enumerate(st.session_state.registre_heures):
                r_idx = start_row + idx
                ws.cell(row=r_idx, column=1, value=data["Date"])
                ws.cell(row=r_idx, column=2, value=data["Horaire"])
                if data["Semaine"]: ws.cell(row=r_idx, column=3, value=data["Semaine"])
                if data["Dimanche"]: ws.cell(row=r_idx, column=4, value=data["Dimanche"])
                if data["Nuit"]: ws.cell(row=r_idx, column=5, value=data["Nuit"])
                ws.cell(row=r_idx, column=6, value=data["Motif"])

            output_filename = f"DEMANDE_HS_{agent_nom.replace(' ', '_')}_{mois_annee.replace(' ', '_')}.xlsx"
            wb.save(output_filename)

            with open(output_filename, "rb") as f:
                st.download_button(
                    label="💾 Télécharger l'Excel signé pour la hiérarchie",
                    data=f,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"Erreur lors de la génération du document : {e}")
